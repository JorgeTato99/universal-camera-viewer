"""
Router para endpoints de métricas de publicación MediaMTX.

Proporciona APIs para consultar, analizar y exportar métricas
de streaming en tiempo real e históricas.
"""

import logging
from fastapi import APIRouter, HTTPException, Depends, Query, status, BackgroundTasks
from fastapi.responses import StreamingResponse
from typing import Optional, Dict, Any
from datetime import datetime
import csv
import json
import io
from pathlib import Path
import uuid

from api.schemas.requests.mediamtx_requests import (
    GetMetricsRequest, ExportMetricsRequest, MetricTimeRange
)
from api.schemas.responses.mediamtx_responses import (
    PublicationMetricsResponse, MetricsExportResponse,
    GlobalMetricsSummaryResponse, PublishMetricsSnapshot,
    PaginatedMetricsResponse
)
from api.dependencies import create_response
from services.database.mediamtx_db_service import get_mediamtx_db_service
from services.camera_manager_service import camera_manager_service
from utils.exceptions import ServiceError
from api.validators.mediamtx_validators import (
    validate_metric_time_range, validate_export_format
)


logger = logging.getLogger(__name__)


router = APIRouter(
    prefix="/api/publishing/metrics",
    tags=["publishing-metrics"],
    responses={404: {"description": "Not found"}},
)


@router.get(
    "/current/{camera_id}",
    response_model=PublishMetricsSnapshot,
    summary="Obtener métricas actuales",
    description="""
    Obtiene la métrica más reciente de una cámara en publicación.
    
    Este endpoint está optimizado para dashboards en tiempo real
    y retorna un objeto simple compatible con el frontend.
    """
)
async def get_current_metrics(camera_id: str) -> PublishMetricsSnapshot:
    """
    Obtiene las métricas más recientes de una cámara.
    
    Args:
        camera_id: Identificador único de la cámara
        
    Returns:
        PublishMetricsSnapshot con la métrica más reciente
        
    Raises:
        HTTPException: Si la cámara no existe o no está publicando
    """
    logger.debug(f"Obteniendo métricas actuales para cámara {camera_id}")
    
    try:
        # Validar que la cámara existe
        camera = await camera_manager_service.get_camera(camera_id)
        if not camera:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Cámara {camera_id} no encontrada"
            )
        
        # Obtener servicio de BD
        db_service = get_mediamtx_db_service()
        await db_service.initialize()
        
        # Intentar obtener métricas del servicio en tiempo real primero
        metric_snapshot = None
        
        try:
            from services.mediamtx_metrics_service import get_mediamtx_metrics_service
            metrics_service = get_mediamtx_metrics_service()
            
            current_metrics = metrics_service.get_current_metrics(camera_id)
            if current_metrics:
                # Calcular quality score
                quality_score = metrics_service.calculate_quality_score(current_metrics)
                
                # Crear snapshot desde métricas en tiempo real
                metric_snapshot = PublishMetricsSnapshot(
                    camera_id=camera_id,
                    timestamp=current_metrics.timestamp.isoformat(),
                    fps=current_metrics.fps,
                    bitrate_kbps=current_metrics.bitrate_kbps,
                    viewers=current_metrics.viewer_count,
                    frames_sent=current_metrics.frames,
                    bytes_sent=int(current_metrics.size_kb * 1024),
                    quality_score=quality_score,
                    status=_determine_stream_status(quality_score)
                )
                logger.debug(f"Métricas obtenidas del servicio en tiempo real para {camera_id}")
        except Exception as e:
            logger.warning(f"No se pudieron obtener métricas del servicio: {e}")
        
        # Si no hay métricas en tiempo real, obtener de la BD
        if not metric_snapshot:
            latest_metric = await db_service.get_latest_publication_metric(camera_id)
            
            if not latest_metric:
                logger.warning(f"Cámara {camera_id} no tiene métricas recientes")
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail={
                        "error": "No hay métricas disponibles para esta cámara",
                        "camera_id": camera_id
                    }
                )
            
            # Mapear campos del backend al formato esperado por el frontend
            metric_snapshot = PublishMetricsSnapshot(
                camera_id=camera_id,
                timestamp=latest_metric['timestamp'].isoformat() if latest_metric.get('timestamp') else datetime.utcnow().isoformat(),
                fps=latest_metric.get('fps', 0.0),
                bitrate_kbps=latest_metric.get('bitrate_kbps', 0.0),
                viewers=latest_metric.get('viewer_count', 0),  # Mapeo: viewer_count → viewers
                frames_sent=latest_metric.get('frames', 0),    # Mapeo: frames → frames_sent
                bytes_sent=int(latest_metric.get('size_kb', 0) * 1024),  # Conversión: KB → bytes
                quality_score=latest_metric.get('quality_score'),
                status=_determine_stream_status(latest_metric.get('quality_score'))
            )
        
        logger.info(f"Métricas actuales obtenidas para {camera_id}: FPS={metric_snapshot.fps}, viewers={metric_snapshot.viewers}")
        
        return metric_snapshot
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error obteniendo métricas actuales para {camera_id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error obteniendo métricas actuales"
        )


@router.get(
    "/{camera_id}",
    response_model=PublicationMetricsResponse,
    summary="Obtener métricas de publicación",
    description="""
    Obtiene métricas detalladas de publicación para una cámara específica.
    
    Incluye:
    - FPS, bitrate, frames transmitidos
    - Calidad de stream y uso de recursos
    - Estadísticas de viewers (opcional)
    - Resumen estadístico del período
    """
)
async def get_camera_metrics(
    camera_id: str,
    time_range: MetricTimeRange = Query(
        MetricTimeRange.LAST_HOUR,
        description="Rango de tiempo predefinido"
    ),
    start_time: Optional[datetime] = Query(
        None,
        description="Tiempo de inicio para rango CUSTOM"
    ),
    end_time: Optional[datetime] = Query(
        None,
        description="Tiempo de fin para rango CUSTOM"
    ),
    include_viewers: bool = Query(
        False,
        description="Incluir estadísticas de viewers"
    ),
    aggregate_interval: Optional[str] = Query(
        None,
        regex="^(1m|5m|15m|1h|1d)$",
        description="Intervalo de agregación de datos"
    )
) -> PublicationMetricsResponse:
    """
    Obtiene métricas de publicación de una cámara.
    
    Args:
        camera_id: Identificador único de la cámara
        time_range: Rango de tiempo para las métricas
        start_time: Inicio del rango personalizado
        end_time: Fin del rango personalizado
        include_viewers: Si incluir estadísticas de viewers
        aggregate_interval: Intervalo para agregar datos
        
    Returns:
        PublicationMetricsResponse con métricas y resumen
        
    Raises:
        HTTPException: Si la cámara no existe o no está publicando
    """
    logger.debug(f"Obteniendo métricas para cámara {camera_id}, rango: {time_range}")
    
    try:
        # Validar que la cámara existe
        camera = await camera_manager_service.get_camera(camera_id)
        if not camera:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Cámara {camera_id} no encontrada"
            )
        
        # Validar rango de tiempo personalizado
        if time_range == MetricTimeRange.CUSTOM:
            if not start_time or not end_time:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="start_time y end_time son requeridos para rango CUSTOM"
                )
            start_time, end_time = validate_metric_time_range(start_time, end_time)
        
        # Crear request object
        request = GetMetricsRequest(
            time_range=time_range,
            start_time=start_time,
            end_time=end_time,
            include_viewers=include_viewers,
            aggregate_interval=aggregate_interval
        )
        
        # Obtener servicio de BD
        db_service = get_mediamtx_db_service()
        await db_service.initialize()
        
        # Consultar métricas
        metrics_data = await db_service.get_publication_metrics(camera_id, request)
        
        if not metrics_data['publication_id']:
            logger.warning(f"Cámara {camera_id} no tiene publicación activa")
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={
                    "error": "La cámara no está publicando actualmente",
                    "camera_id": camera_id
                }
            )
        
        # TODO: Si se especifica aggregate_interval, agregar los data points
        if aggregate_interval:
            metrics_data['data_points'] = _aggregate_metrics(
                metrics_data['data_points'],
                aggregate_interval
            )
        
        # Construir respuesta
        response = PublicationMetricsResponse(**metrics_data)
        
        logger.info(f"Métricas obtenidas para {camera_id}: "
                   f"{len(response.data_points)} puntos de datos")
        
        return response
        
    except HTTPException:
        raise
    except ServiceError as e:
        logger.error(f"ServiceError obteniendo métricas: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )
    except Exception as e:
        logger.exception(f"Error inesperado obteniendo métricas para {camera_id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error obteniendo métricas"
        )


@router.get(
    "/{camera_id}/history",
    response_model=PaginatedMetricsResponse,
    summary="Historial de métricas paginado",
    description="Obtiene el historial de métricas con paginación para gráficos temporales"
)
async def get_metrics_history(
    camera_id: str,
    time_range: MetricTimeRange = Query(
        MetricTimeRange.LAST_HOUR,
        description="Rango de tiempo predefinido"
    ),
    start_time: Optional[datetime] = Query(None),
    end_time: Optional[datetime] = Query(None),
    page: int = Query(1, ge=1, description="Número de página"),
    page_size: int = Query(100, ge=10, le=1000, description="Tamaño de página"),
    min_quality_score: Optional[float] = Query(
        None,
        ge=0,
        le=100,
        description="Score de calidad mínimo"
    ),
    max_error_rate: Optional[float] = Query(
        None,
        ge=0,
        le=100,
        description="Tasa de error máxima %"
    )
) -> PaginatedMetricsResponse:
    """
    Obtiene historial paginado de métricas para gráficos temporales.
    
    Retorna las métricas en formato simplificado compatible con el frontend.
    """
    logger.debug(f"Obteniendo historial de métricas para {camera_id}, página {page}")
    
    try:
        # Validar que la cámara existe
        camera = await camera_manager_service.get_camera(camera_id)
        if not camera:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Cámara {camera_id} no encontrada"
            )
        
        # Validar rango de tiempo personalizado
        if time_range == MetricTimeRange.CUSTOM:
            if not start_time or not end_time:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="start_time y end_time son requeridos para rango CUSTOM"
                )
            start_time, end_time = validate_metric_time_range(start_time, end_time)
        
        # Crear request object
        request = GetMetricsRequest(
            time_range=time_range,
            start_time=start_time,
            end_time=end_time,
            include_viewers=False,
            aggregate_interval=None
        )
        
        # Obtener servicio de BD
        db_service = get_mediamtx_db_service()
        await db_service.initialize()
        
        # Obtener métricas con paginación
        metrics_data = await db_service.get_publication_metrics(camera_id, request)
        
        if not metrics_data['publication_id']:
            logger.warning(f"Cámara {camera_id} no tiene publicación activa")
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={
                    "error": "La cámara no está publicando actualmente",
                    "camera_id": camera_id
                }
            )
        
        # Filtrar datos si es necesario
        data_points = metrics_data['data_points']
        if min_quality_score is not None or max_error_rate is not None:
            filtered_points = []
            for point in data_points:
                # Filtrar por calidad
                if min_quality_score and point.get('quality_score'):
                    if point['quality_score'] < min_quality_score:
                        continue
                
                # Filtrar por tasa de error
                if max_error_rate and point.get('frames') and point.get('dropped_frames'):
                    error_rate = (point['dropped_frames'] / point['frames']) * 100
                    if error_rate > max_error_rate:
                        continue
                
                filtered_points.append(point)
            data_points = filtered_points
        
        # Paginar resultados
        total = len(data_points)
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_points = data_points[start_idx:end_idx]
        
        # Convertir a PublishMetricsSnapshot
        metrics_snapshots = []
        for point in paginated_points:
            snapshot = PublishMetricsSnapshot(
                camera_id=camera_id,
                timestamp=point.get('timestamp').isoformat() if point.get('timestamp') else "",
                fps=point.get('fps', 0.0),
                bitrate_kbps=point.get('bitrate_kbps', 0.0),
                viewers=point.get('viewer_count', 0),
                frames_sent=point.get('frames', 0),
                bytes_sent=int(point.get('size_kb', 0) * 1024) if point.get('size_kb') else 0,
                quality_score=point.get('quality_score'),
                status=_determine_stream_status(point.get('quality_score'))
            )
            metrics_snapshots.append(snapshot)
        
        # Determinar rango de tiempo
        time_range_dict = {
            "start": start_time.isoformat() if start_time else (data_points[0]['timestamp'].isoformat() if data_points else ""),
            "end": end_time.isoformat() if end_time else (data_points[-1]['timestamp'].isoformat() if data_points else "")
        }
        
        response = PaginatedMetricsResponse(
            camera_id=camera_id,
            total=total,
            page=page,
            page_size=page_size,
            metrics=metrics_snapshots,
            time_range=time_range_dict
        )
        
        logger.info(f"Historial obtenido para {camera_id}: {len(metrics_snapshots)} métricas, página {page}/{(total + page_size - 1) // page_size}")
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error obteniendo historial de métricas para {camera_id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error obteniendo historial de métricas"
        )


@router.post(
    "/{camera_id}/export",
    response_model=MetricsExportResponse,
    summary="Exportar métricas",
    description="Exporta métricas a archivo descargable (CSV, JSON, Excel)"
)
async def export_camera_metrics(
    camera_id: str,
    request: ExportMetricsRequest,
    background_tasks: BackgroundTasks
) -> MetricsExportResponse:
    """
    Exporta métricas de una cámara a archivo.
    
    Args:
        camera_id: ID de la cámara
        request: Parámetros de exportación
        background_tasks: Para procesar exportación en background
        
    Returns:
        MetricsExportResponse con URL de descarga
    """
    logger.info(f"Exportando métricas para {camera_id} en formato {request.format}")
    
    try:
        # Validar formato
        export_format = validate_export_format(request.format)
        
        # Generar ID único para la exportación
        export_id = str(uuid.uuid4())
        
        # Definir ruta del archivo
        export_dir = Path("exports/metrics")
        export_dir.mkdir(parents=True, exist_ok=True)
        
        extension_map = {
            'csv': 'csv',
            'json': 'json',
            'excel': 'xlsx'
        }
        
        filename = f"metrics_{camera_id}_{export_id}.{extension_map[export_format]}"
        file_path = export_dir / filename
        
        # Programar generación del archivo en background
        # NOTA: La generación real del archivo está implementada pero
        # puede fallar si no hay datos suficientes en la BD
        background_tasks.add_task(
            _generate_export_file,
            camera_id=camera_id,
            file_path=file_path,
            format=export_format,
            time_range=request.time_range,
            include_raw=request.include_raw_data
        )
        
        # URL de descarga (relativa al servidor)
        download_url = f"/api/publishing/metrics/download/{export_id}"
        
        # Tiempo de expiración (24 horas)
        expires_at = datetime.utcnow().replace(hour=0, minute=0, second=0)
        expires_at = expires_at.replace(day=expires_at.day + 1)
        
        return MetricsExportResponse(
            export_id=export_id,
            format=export_format,
            file_size_bytes=0,  # Se actualizará cuando se genere
            download_url=download_url,
            expires_at=expires_at,
            record_count=0  # Se actualizará cuando se genere
        )
        
    except Exception as e:
        logger.exception(f"Error iniciando exportación para {camera_id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error iniciando exportación"
        )


@router.get(
    "/download/{export_id}",
    summary="Descargar archivo exportado",
    description="Descarga un archivo de métricas previamente exportado"
)
async def download_export(export_id: str):
    """
    Descarga un archivo de exportación por su ID.
    
    Args:
        export_id: ID único de la exportación
        
    Returns:
        StreamingResponse con el archivo
        
    Raises:
        HTTPException: Si el archivo no existe o expiró
    """
    try:
        # Buscar archivo en directorio de exports
        export_dir = Path("exports/metrics")
        
        # Buscar archivo con el export_id
        matching_files = list(export_dir.glob(f"*_{export_id}.*"))
        
        if not matching_files:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Archivo de exportación no encontrado o expirado"
            )
        
        file_path = matching_files[0]
        
        # Determinar content type
        content_type_map = {
            '.csv': 'text/csv',
            '.json': 'application/json',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        content_type = content_type_map.get(
            file_path.suffix,
            'application/octet-stream'
        )
        
        # Retornar archivo
        return StreamingResponse(
            file_path.open('rb'),
            media_type=content_type,
            headers={
                "Content-Disposition": f"attachment; filename={file_path.name}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error descargando export {export_id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error descargando archivo"
        )


@router.get(
    "/statistics/summary",
    response_model=GlobalMetricsSummaryResponse,
    summary="Resumen global de métricas",
    description="Dashboard con estadísticas globales del sistema de publicación"
)
async def get_global_statistics() -> GlobalMetricsSummaryResponse:
    """
    Obtiene un resumen global de métricas del sistema.
    
    Incluye:
    - Total de cámaras publicando
    - Viewers activos totales
    - Ancho de banda agregado
    - Top cámaras por audiencia
    - Alertas activas
    
    Returns:
        GlobalMetricsSummaryResponse con dashboard de métricas
    """
    logger.debug("Obteniendo estadísticas globales del sistema")
    
    try:
        db_service = get_mediamtx_db_service()
        await db_service.initialize()
        
        # Obtener resumen global de la BD
        summary_data = await db_service.get_global_metrics_summary()
        
        # Enriquecer con datos del servicio de métricas si está disponible
        try:
            from services.mediamtx_metrics_service import get_mediamtx_metrics_service
            metrics_service = get_mediamtx_metrics_service()
            
            # Actualizar con métricas en tiempo real si hay publicaciones activas
            if summary_data.get('total_cameras_publishing', 0) > 0:
                # Obtener quality score promedio de cámaras activas
                active_scores = []
                for camera_info in summary_data.get('top_cameras', []):
                    camera_id = camera_info.get('camera_id')
                    if camera_id:
                        current_metrics = metrics_service.get_current_metrics(camera_id)
                        if current_metrics:
                            quality_score = metrics_service.calculate_quality_score(current_metrics)
                            active_scores.append(quality_score)
                
                if active_scores:
                    summary_data['avg_quality_score'] = sum(active_scores) / len(active_scores)
                    
        except Exception as e:
            logger.warning(f"No se pudieron obtener métricas en tiempo real: {e}")
        
        # Construir respuesta
        response = GlobalMetricsSummaryResponse(**summary_data)
        
        logger.info(f"Estadísticas globales: {response.total_cameras_publishing} cámaras, "
                   f"{response.total_active_viewers} viewers")
        
        return response
        
    except Exception as e:
        logger.exception("Error obteniendo estadísticas globales")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error obteniendo estadísticas"
        )


# === Funciones auxiliares ===

def _determine_stream_status(quality_score: Optional[float]) -> Optional[str]:
    """
    Determina el estado del stream basado en el quality score.
    
    Args:
        quality_score: Score de calidad (0-100)
        
    Returns:
        Estado: 'optimal', 'degraded', 'poor' o None
    """
    if quality_score is None:
        return None
        
    if quality_score >= 80:
        return "optimal"
    elif quality_score >= 50:
        return "degraded"
    else:
        return "poor"


def _aggregate_metrics(
    data_points: list,
    interval: str
) -> list:
    """
    Agrega puntos de métricas según el intervalo especificado.
    
    Args:
        data_points: Lista de puntos de métrica
        interval: Intervalo de agregación (1m, 5m, etc.)
        
    Returns:
        Lista de puntos agregados
    """
    # TODO: Implementar agregación real
    # Por ahora retornamos los datos sin modificar
    # La agregación debería promediar valores dentro de cada intervalo
    logger.warning(f"Agregación con intervalo {interval} no implementada aún")
    return data_points


async def _generate_export_file(
    camera_id: str,
    file_path: Path,
    format: str,
    time_range: MetricTimeRange,
    include_raw: bool
):
    """
    Genera el archivo de exportación en background.
    
    Args:
        camera_id: ID de la cámara
        file_path: Ruta donde guardar el archivo
        format: Formato de exportación
        time_range: Rango de tiempo
        include_raw: Si incluir datos crudos
    """
    try:
        logger.info(f"Generando archivo de exportación: {file_path}")
        
        # Obtener datos
        db_service = get_mediamtx_db_service()
        await db_service.initialize()
        
        request = GetMetricsRequest(
            time_range=time_range,
            include_viewers=True
        )
        
        metrics_data = await db_service.get_publication_metrics(camera_id, request)
        
        # Generar archivo según formato
        if format == 'csv':
            await _export_to_csv(metrics_data, file_path, include_raw)
        elif format == 'json':
            await _export_to_json(metrics_data, file_path, include_raw)
        elif format == 'excel':
            await _export_to_excel(metrics_data, file_path, include_raw)
        
        logger.info(f"Archivo de exportación generado: {file_path}")
        
    except Exception as e:
        logger.exception(f"Error generando exportación: {e}")
        # Eliminar archivo parcial si existe
        if file_path.exists():
            file_path.unlink()


async def _export_to_csv(data: dict, file_path: Path, include_raw: bool):
    """Exporta métricas a CSV."""
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        # Definir columnas
        fieldnames = [
            'timestamp', 'fps', 'bitrate_kbps', 'frames', 
            'dropped_frames', 'quality_score', 'viewer_count',
            'cpu_usage_percent', 'memory_usage_mb'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        # Escribir datos
        for point in data['data_points']:
            row = {
                'timestamp': point.get('timestamp'),
                'fps': point.get('fps'),
                'bitrate_kbps': point.get('bitrate_kbps'),
                'frames': point.get('frames'),
                'dropped_frames': point.get('dropped_frames'),
                'quality_score': point.get('quality_score'),
                'viewer_count': point.get('viewer_count'),
                'cpu_usage_percent': point.get('cpu_usage_percent'),
                'memory_usage_mb': point.get('memory_usage_mb')
            }
            writer.writerow(row)
        
        # Si no incluir datos crudos, agregar resumen al final
        if not include_raw and data.get('summary'):
            writer.writerow({})  # Línea vacía
            writer.writerow({'timestamp': 'SUMMARY'})
            writer.writerow({
                'timestamp': 'Average',
                'fps': data['summary'].get('avg_fps'),
                'bitrate_kbps': data['summary'].get('avg_bitrate_kbps'),
                'quality_score': data['summary'].get('avg_quality_score'),
                'viewer_count': data['summary'].get('avg_viewers')
            })


async def _export_to_json(data: dict, file_path: Path, include_raw: bool):
    """Exporta métricas a JSON."""
    export_data = {
        'camera_id': data['camera_id'],
        'publication_id': data['publication_id'],
        'time_range': data['time_range'],
        'export_date': datetime.utcnow().isoformat(),
        'summary': data['summary']
    }
    
    if include_raw:
        export_data['raw_data'] = data['data_points']
    
    if data.get('viewer_stats'):
        export_data['viewer_stats'] = data['viewer_stats']
    
    with open(file_path, 'w', encoding='utf-8') as jsonfile:
        json.dump(export_data, jsonfile, indent=2, default=str)


async def _export_to_excel(data: dict, file_path: Path, include_raw: bool):
    """
    Exporta métricas a Excel.
    
    NOTA: Requiere que openpyxl esté instalado. Si no está disponible,
    se exportará como CSV como alternativa.
    """
    # Requiere openpyxl
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Hoja de resumen
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Escribir resumen
        summary_data = [
            ["Metric", "Value"],
            ["Camera ID", data['camera_id']],
            ["Time Range", data['time_range']],
            ["Average FPS", data['summary'].get('avg_fps', 0)],
            ["Average Bitrate (kbps)", data['summary'].get('avg_bitrate_kbps', 0)],
            ["Total Frames", data['summary'].get('total_frames', 0)],
            ["Average Quality Score", data['summary'].get('avg_quality_score', 0)],
            ["Peak Viewers", data['summary'].get('peak_viewers', 0)],
            ["Average Viewers", data['summary'].get('avg_viewers', 0)]
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Hoja de datos crudos si se solicita
        if include_raw and data['data_points']:
            ws_raw = wb.create_sheet("Raw Data")
            
            # Headers
            headers = [
                'Timestamp', 'FPS', 'Bitrate (kbps)', 'Frames',
                'Dropped Frames', 'Quality Score', 'Viewer Count',
                'CPU %', 'Memory (MB)'
            ]
            ws_raw.append(headers)
            
            # Datos
            for point in data['data_points']:
                row = [
                    point.get('timestamp'),
                    point.get('fps'),
                    point.get('bitrate_kbps'),
                    point.get('frames'),
                    point.get('dropped_frames'),
                    point.get('quality_score'),
                    point.get('viewer_count'),
                    point.get('cpu_usage_percent'),
                    point.get('memory_usage_mb')
                ]
                ws_raw.append(row)
            
            # Ajustar anchos de columna
            for col_idx, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                ws_raw.column_dimensions[col_letter].width = len(header) + 5
        
        # Guardar archivo
        wb.save(file_path)
        
    except ImportError:
        # Si no está instalado openpyxl, generar CSV como fallback
        logger.warning("openpyxl no disponible, exportando como CSV")
        await _export_to_csv(data, file_path.with_suffix('.csv'), include_raw)